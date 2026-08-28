import openpyxl
import json

filepath = r"C:\Users\k0115\OneDrive\바탕 화면\2026-07월매입매출현황의 복사본.xlsx"
wb = openpyxl.load_workbook(filepath, data_only=True)
print("Sheet names:", wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\n==========================================")
    print(f"=== Sheet: {sheet} (rows={ws.max_row}, cols={ws.max_column}) ===")
    print(f"==========================================")
    for r in range(1, min(15, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(15, ws.max_column + 1))]
        if any(v is not None for v in row_vals):
            print(f"Row {r:2d}:", [str(v) if v is not None else "" for v in row_vals])
